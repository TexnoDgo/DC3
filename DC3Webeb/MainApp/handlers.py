import fitz
import os
from os.path import basename
from openpyxl import Workbook
import qrcode
import zipfile


from .models import *

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


# Функция конвертации PDF в PNG
def convert_pdf_to_png(file):
    doc = fitz.open(file.path)
    page = doc.loadPage(0)
    pix = page.getPixmap()
    pdf_file_name = str(file)
    png_file_name = '{}{}'.format(pdf_file_name[19:-3], 'png')
    png_full_path = os.path.join(BASE_DIR, 'media/COMPONENT_PNG_DRAW/') + png_file_name
    print(png_full_path)
    output = png_full_path
    pix.writePNG(output)
    return png_full_path


# Функция генерации QR-кодов для позиции
def generative_qr_code(code, qr_type):
    if qr_type == 'SMART':
        # TODO: Ввести корректный адрес хоста
        detail_url = 'host address' + 'path' + code
        img = qrcode.make(detail_url)
        img_path = os.path.join(BASE_DIR, 'media/SMART_PROD_QR_CODE/') + code + '.png'
        img.save(img_path)
        img_name = 'SMART_PROD_QR_CODE/' + code + '.png'
    elif qr_type == 'PK':
        detail_url = code
        img = qrcode.make(detail_url)
        img_path = os.path.join(BASE_DIR, 'media/PK_PROD_QR_CODE/') + code + '.png'
        img.save(img_path)
        img_name = 'PK_PROD_QR_CODE/' + code + '.png'

    return img_name


# Функция распознования формата чертежа
def file_size_check(file):
    # TODO: Добавить проверку файлов А1, А0. А так же их вертикальные исполнения
    try:
        doc = fitz.open(file)
    except:
        print(file)
        print('Ошибка чтения файла')

    page = doc[0]
    long = page.MediaBox[2]
    height = page.MediaBox[3]
    if 550 < long < 700:
        if 800 < height < 900:
            return "A4"
    elif 1150 < long < 1250:
        if 800 < height < 900:
            return "A3"
    elif 1600 < long < 1750:
        if 1100 < height < 1300:
            return "A2"


# Функция создание модифицированного рабочего чертежа
def create_work_pdf(old_pdf_file, new_pdf_file, qr_code_path, qr_code_path2, project_name, order_name, mather_assembly_name, quantity,
                    file_size):
    print("Выполняется создание рабочаего чертежа")

    real_old_pdf_file = os.path.join(BASE_DIR, 'media/') + str(old_pdf_file)
    real_old_pdf_file = real_old_pdf_file.replace('\\', '/')
    real_new_pdf_file = os.path.join(BASE_DIR, 'media/') + str(new_pdf_file)
    real_new_pdf_file = real_new_pdf_file.replace('\\', '/')
    real_qr_code_path = os.path.join(BASE_DIR, 'media/') + str(qr_code_path)
    real_qr_code_path = real_qr_code_path.replace('\\', '/')
    real_qr_code_path2 = os.path.join(BASE_DIR, 'media/') + str(qr_code_path2)
    real_qr_code_path2 = real_qr_code_path2.replace('\\', '/')
    doc = fitz.open(real_old_pdf_file)

    mstb = 1.378
    print(file_size)
    if file_size == "A4":
        qr_rect =               fitz.Rect((666-64) / mstb, 1000 / mstb, ((666-64) + 64) / mstb,  (1000 + 64) / mstb)
        qr_rect2 =              fitz.Rect((666) / mstb, 1000 / mstb, ((666) + 64) / mstb,  (1000 + 64) / mstb)
        project_rect =          fitz.Rect((609-50) / mstb, (1082-230) / mstb, (609 + 200) / mstb, (1082 + 16) / mstb)
        order_rect =            fitz.Rect((609-50) / mstb, (1097-230) / mstb, (609 + 200) / mstb, (1097 + 16) / mstb)
        mather_assembly_rect =  fitz.Rect((609-50) / mstb, (1111-230) / mstb, (609 + 200) / mstb, (1111 + 16) / mstb)
        quantity_rect =         fitz.Rect((609-50) / mstb, (1126-230) / mstb, (609 + 200) / mstb, (1126 + 16) / mstb)
    elif file_size == 'A3':
        i = 595 * mstb
        qr_rect =               fitz.Rect((666-64 + i) / mstb, 1000 / mstb, ((666-64 + i) +  64) / mstb, (1000 + 64) / mstb)
        qr_rect2 =              fitz.Rect((666 + i) / mstb, 1000 / mstb, ((666 + i) +  64) / mstb, (1000 + 64) / mstb)
        project_rect =          fitz.Rect(((609-50) + i) / mstb, (1082-230) / mstb, ((609 + i) + 200) / mstb, (1082 + 16) / mstb)
        order_rect =            fitz.Rect(((609-50) + i) / mstb, (1097-230) / mstb, ((609 + i) + 200) / mstb, (1097 + 16) / mstb)
        mather_assembly_rect =  fitz.Rect(((609-50) + i) / mstb, (1111-230) / mstb, ((609 + i) + 200) / mstb, (1111 + 16) / mstb)
        quantity_rect =         fitz.Rect(((609-50) + i) / mstb, (1126-230) / mstb, ((609 + i) + 200) / mstb, (1126 + 16) / mstb)
    elif file_size == 'A2':
        i = 1089 * mstb
        j = 348 * mstb
        qr_rect =               fitz.Rect((666 + i) / mstb, (1000 + j) / mstb, ((666 + i) + 64) / mstb,
                                            ((1000 + j) + 64) / mstb)
        qr_rect2 = fitz.Rect((666-64 + i) / mstb, (1000 + j) / mstb, ((666-64 + i) + 64) / mstb,
                            ((1000 + j) + 64) / mstb)
        project_rect =          fitz.Rect((609-50 + i) / mstb, (1082-230 + j) / mstb, ((609 + i) + 200) / mstb,
                                            ((1082 + j) + 16) / mstb)
        order_rect =            fitz.Rect((609-50 + i) / mstb, (1097-230 + j) / mstb, ((609 + i) + 200) / mstb,
                                            ((1097 + j) + 16) / mstb)
        mather_assembly_rect =  fitz.Rect((609-50 + i) / mstb, (1111-230 + j) / mstb, ((609 + i) + 200) / mstb,
                                            ((1111 + j) + 16) / mstb)
        quantity_rect =         fitz.Rect((609-50 + i) / mstb, (1126-230 + j) / mstb, ((609 + i) + 200) / mstb,
                                            ((1126 + j) + 16) / mstb)

    project_text = "Проект: " + project_name
    order_text = "Заказ: " + order_name
    assembly_text = "Сборка: " + mather_assembly_name.info.title
    quantity_text = "Кол-во: " + str(quantity)
    page = doc[0]
    page.insertImage(qr_rect, filename=real_qr_code_path)
    page.insertImage(qr_rect2, filename=real_qr_code_path2)
    # TODO: поместить файл шрифта
    font_path = os.path.join(BASE_DIR, 'static/admin/fonts/20289.ttf')
    print("Путь к файлу шрифта: " + font_path)

    rc1 = page.insertTextbox(project_rect, project_text, fontsize=9,
                             fontname="FreeSans",  # a PDF standard font
                             fontfile=font_path,
                             # could be a file on your system
                             align=0)
    rc2 = page.insertTextbox(order_rect, order_text, fontsize=9,
                             fontname="FreeSans",  # a PDF standard font
                             fontfile=font_path,
                             # could be a file on your system
                             align=0)
    rc3 = page.insertTextbox(mather_assembly_rect, assembly_text, fontsize=9,
                             fontname="FreeSans",  # a PDF standard font
                             fontfile=font_path,
                             # could be a file on your system
                             align=0)
    rc4 = page.insertTextbox(quantity_rect, quantity_text, fontsize=9,
                             fontname="FreeSans",  # a PDF standard font
                             fontfile=font_path,
                             # could be a file on your system
                             align=0)

    doc.save(real_new_pdf_file)
    print("Рабочий чертеж создан!")

    return new_pdf_file


# Функция подготовка к модификации PDF-чертежа
def create_modify_draw(component, position, order, code):
    try:
        # Получение формата pdf чертежа
        pdf_path = component.pdf_draw.path
        component_pdf_size = file_size_check(pdf_path)
        # Получение пути к модицицированному черетжу с qr кодом
        sticker_draw_pdf_path = 'PDF_DRAW_WT_QR_CODE/' + code + '.PDF'

        # Создание модифицированного pdf чертежа с qr кодом
        sticker_draw_pdf = create_work_pdf(component.pdf_draw, sticker_draw_pdf_path, position.pk_prod_qr_code,
                                           position.smart_prod_qr_code, order.project.info.title, order.info.title,
                                           position.m_assembly, position.quantity, component_pdf_size)
        return sticker_draw_pdf
    except Exception:
        return False


# Функция формирования таблица спецификации
def specification_table(m_position, project):
    # Создание таблицы Excel
    wb = Workbook()
    ws = wb.active
    # Заполнение полей проекта. Шапка
    ws['A1'] = 'Проект'
    ws['B1'] = 'Устройство'
    ws['C1'] = 'Сборка'
    # Заполнение полей проекта. Информация
    ws['A2'] = project.info.title
    ws['B2'] = project.device.info.title
    ws['C2'] = m_position.m_assembly.info.title
    # Заполнение полей проекта. Поля позиций
    ws['A3'] = '#'
    ws['B3'] = 'Наименование'
    ws['C3'] = 'Заказ'
    ws['D3'] = 'Кол-во,шт.'
    ws['E3'] = 'Приоритет'
    ws['F3'] = 'Место Хранение'
    # TODO: Реализовать функцию архивации таблицы и чертежей - DONE!
    # Создание архива спецификации сборки
    archive_name = 'Спецификация_Проекта_' + str(project.info.title) + '_Сборка_' + str(m_position.component.info.title)
    archive_file_path = os.path.join(BASE_DIR, 'media/temp/') + "{}.zip".format(archive_name)
    archive_file_path = archive_file_path.replace('\\', '/')
    archive = zipfile.ZipFile(archive_file_path, 'w')
    # Получение значений
    orders = Order.objects.filter(project=project)
    # Счетчик компонентов
    j = 1
    # Начало спецификации
    i = 4
    # Перебор позиций в сборке проекта
    for order in orders:
        positions = Position.objects.filter(order=order)
        for position in positions:
            if position.m_assembly == m_position.component:
                # Внесение данных в поля таблицы
                ws.cell(row=i, column=1, value=j)
                ws.cell(row=i, column=2, value=position.component.info.title)
                ws.cell(row=i, column=3, value=position.order.info.title)
                ws.cell(row=i, column=4, value=position.quantity)
                ws.cell(row=i, column=5, value=position.priority)
                ws.cell(row=i, column=6, value=position.storage.title)
                # Запись чертежа позиции в архив
                archive.write(position.pdf_draw_wt_qr_code.path, basename(position.pdf_draw_wt_qr_code.path))
                j += 1
                i += 1

    table_name = 'project_spec_#_' + str(project.pk) + '_assembly_' + str(m_position.component.pk)
    table_path = os.path.join(BASE_DIR, 'media/temp/') + "{}.xlsx".format(table_name)
    wb.save(table_path)
    # Запись таблицу спецификации в архив
    archive.write(table_path, basename(table_path))
    # Сохранить и закрыть архив

    return archive_file_path.replace('/', '\\')
